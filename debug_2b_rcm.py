
import openpyxl
import os

file_path = r"c:\Users\manum\.gemini\antigravity\gst\032023_32AAMFM4610Q1Z0_GSTR2BQ_05012026.xlsx"

try:
    print(f"Loading workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheet Names: {wb.sheetnames}")
    
    if "ITC Available" in wb.sheetnames:
        ws = wb["ITC Available"]
        print("\n--- Rows Inspection ---")
        
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            # Mimic the logic in analyzer
            row_text_parts = [str(x).lower().strip() for x in row if isinstance(x, str)]
            row_text = " ".join(row_text_parts).replace(",", "")
            
            print(f"Row {idx}: {row_text}")
            
            if "inward supplies" in row_text and "reverse charge" in row_text:
                print(f"  -> MATCH CANDIDATE")
                if "credit notes" not in row_text and "amendment" not in row_text:
                    print("  -> STRICT MATCH SUCCESS")
                else:
                    print("  -> EXCLUDED due to 'credit notes' or 'amendment'")
            
            if idx > 30: break
    else:
        print("Sheet 'ITC Available' NOT FOUND")
        
    wb.close()

except Exception as e:
    print(f"Error: {e}")
