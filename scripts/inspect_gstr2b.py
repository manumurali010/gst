import pandas as pd
import openpyxl

file_path = r"c:\Users\manum\.gemini\antigravity\scratch\gst\032023_32AAMFM4610Q1Z0_GSTR2BQ_05012026.xlsx"

print(f"--- Inspecting {file_path} ---")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    # 1. Check ITC Available Summary
    if "ITC Available" in wb.sheetnames:
        print("\n--- Sheet: ITC Available (Searching for 'All other ITC') ---")
        ws = wb["ITC Available"]
        found = False
        for row in ws.iter_rows(values_only=True):
            row_str = " ".join([str(x) for x in row if x is not None])
            if "All other ITC" in row_str and "Supplies from registered persons" in row_str:
                print(f"FOUND ROW: {row}")
                found = True
        if not found:
            print("Row 'All other ITC - Supplies from registered persons' NOT FOUND.")
            
    # 2. Check B2B-CDNR
    if "B2B-CDNR" in wb.sheetnames:
        print("\n--- Sheet: B2B-CDNR ---")
        ws = wb["B2B-CDNR"]
        # Print header (Row 1-6 usually headers, data starts around 7)
        # Printing first 10 rows of data
        row_count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 6: continue # Skip likely headers
            row_valid = [str(x) for x in row]
            # Print specifically the Note Type column (Index 3 usually) and Tax columns
            # But let's print the whole row first to be sure
            print(f"Row {i+1}: {row}")
            row_count += 1
            if row_count > 10: break
    else:
        print("\nSheet 'B2B-CDNR' NOT FOUND.")
    
    # 3. Check B2B-CDNRA
    if "B2B-CDNRA" in wb.sheetnames:
        print("\n--- Sheet: B2B-CDNRA ---")
        ws = wb["B2B-CDNRA"]
        row_count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
             if i < 6: continue
             print(f"Row {i+1}: {row}")
             row_count += 1
             if row_count > 10: break

except Exception as e:
    print(f"Error: {e}")
