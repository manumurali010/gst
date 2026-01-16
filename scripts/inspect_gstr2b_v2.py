import pandas as pd
import openpyxl
import sys

# Force UTF-8 for output
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\manum\.gemini\antigravity\scratch\gst\032023_32AAMFM4610Q1Z0_GSTR2BQ_05012026.xlsx"

print(f"--- Inspecting {file_path} ---")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 3. Check B2B-CDNRA
    if "B2B-CDNRA" in wb.sheetnames:
        print("\n--- Sheet: B2B-CDNRA ---")
        ws = wb["B2B-CDNRA"]
        row_count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
             # Print headers too (first 6 rows) to see column names
             if i < 4: continue 
             
             # Clean row for printing
             clean_row = []
             for x in row:
                 s = str(x).replace('\u20b9', 'Rs') # Replace Rupee symbol
                 clean_row.append(s)
                 
             print(f"Row {i+1}: {clean_row}")
             row_count += 1
             if row_count > 15: break
    else:
        print("\nSheet 'B2B-CDNRA' NOT FOUND.")

except Exception as e:
    print(f"Error: {e}")
