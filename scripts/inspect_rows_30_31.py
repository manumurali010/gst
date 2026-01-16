import pandas as pd
import openpyxl
import sys

# Force UTF-8 for output
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\manum\.gemini\antigravity\scratch\gst\032023_32AAMFM4610Q1Z0_GSTR2BQ_05012026.xlsx"

print(f"--- Inspecting Rows 30 & 31 in 'ITC Available' Sheet in {file_path} ---")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if "ITC Available" in wb.sheetnames:
        ws = wb["ITC Available"]
        # Accessing by 1-based index
        rows_to_check = [10, 30, 31]
        
        for r_idx in rows_to_check:
            clean_row = []
            for cell in ws[r_idx]:
                 if cell.value is None:
                     clean_row.append("None")
                 else:
                     clean_row.append(str(cell.value).replace('\u20b9', 'Rs').strip())
            print(f"Row {r_idx}: {clean_row}")
    else:
        print("Sheet 'ITC Available' NOT FOUND.")

except Exception as e:
    print(f"Error: {e}")
