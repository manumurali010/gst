import pandas as pd
import openpyxl
import sys

# Force UTF-8 for output
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\manum\.gemini\antigravity\scratch\gst\032023_32AAMFM4610Q1Z0_GSTR2BQ_05012026.xlsx"

print(f"--- Inspecting Row 12 in {file_path} ---")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    sheets_to_check = ["B2B", "B2B-CDNR", "B2B-CDNRA"]
    
    for sheet in sheets_to_check:
        if sheet in wb.sheetnames:
            print(f"\n--- Sheet: {sheet} ---")
            ws = wb[sheet]
            # row is 1-indexed in openpyxl access usually, or we iterate.
            # User likely implies Excel Row 12.
            # openpyxl ws[12] gives the 12th row.
            
            try:
                row_vals = []
                for cell in ws[12]:
                    row_vals.append(str(cell.value))
                print(f"Row 12 data: {row_vals}")
                
                # Check for "Debit"
                row_str = " ".join(row_vals).lower()
                if "debit" in row_str:
                    print(">>> DEBIT NOTE FOUND IN ROW 12 <<<")
                elif "credit" in row_str:
                    print(">>> CREDIT NOTE FOUND IN ROW 12 <<<")
                else:
                    print(">>> NO NOTE TYPE FOUND IN ROW 12 <<<")
                    
            except Exception as e:
                print(f"Could not read Row 12: {e}")
        else:
             print(f"\n--- Sheet: {sheet} NOT FOUND ---")

except Exception as e:
    print(f"Error: {e}")
