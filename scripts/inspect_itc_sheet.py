import pandas as pd
import openpyxl
import sys

# Force UTF-8 for output
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\manum\.gemini\antigravity\scratch\gst\032023_32AAMFM4610Q1Z0_GSTR2BQ_05012026.xlsx"

print(f"--- Inspecting 'ITC Available' Sheet in {file_path} ---")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if "ITC Available" in wb.sheetnames:
        ws = wb["ITC Available"]
        print("\n--- Sheet: ITC Available (First 20 Rows) ---")
        for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
            # Clean row for printing
            clean_row = []
            for x in row:
                if x is None:
                    clean_row.append("None")
                else:
                    clean_row.append(str(x).replace('\u20b9', 'Rs').strip())
            print(f"Row {i+1}: {clean_row}")
    else:
        print("Sheet 'ITC Available' NOT FOUND.")

except Exception as e:
    print(f"Error: {e}")
