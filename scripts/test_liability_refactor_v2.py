import unittest
from unittest.mock import patch, MagicMock
import os
import sys
import pandas as pd
import openpyxl

# Add src to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from src.services.scrutiny_parser import ScrutinyParser

class TestLiabilityRefactor(unittest.TestCase):
    
    def setUp(self):
        self.parser = ScrutinyParser()
        self.test_excel = "test_liability.xlsx"
        self.test_pdf_3b = "test_3b.pdf"
        self.test_pdf_1 = "test_1.pdf"
        
    def tearDown(self):
        if os.path.exists(self.test_excel):
            os.remove(self.test_excel)
        # PDFs are mocked, so no file cleanup needed for them unless we create empty files to pass os.path.exists
        if os.path.exists(self.test_pdf_3b):
            os.remove(self.test_pdf_3b)
        if os.path.exists(self.test_pdf_1):
            os.remove(self.test_pdf_1)

    def create_mock_excel(self, filename, gstr3b_val=100.0, gstr1_val=120.0):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tax Liability"
        
        # Headers
        ws['B5'] = "Liability per 3B"
        ws['F5'] = "Liability per GSTR-1"
        ws['J5'] = "Difference"
        
        ws['A4'] = "1. Test Issue"
        
        # Row 4 (Headers L0)
        # Cols: 3B=B-E. Ref=F-I. Diff=J-M
        # indices 0..
        # Pandas reads header=[4,5].
        # We need rigorous structure or just enough for parser?
        # Parser checks col names.
        
        for c in range(2, 14):
            ws.cell(row=5, column=c, value="HEAD")
        
        # Actual tax heads in Row 6 (Index 5)
        tax_heads = ["IGST", "CGST", "SGST", "CESS"]
        for i, val in enumerate(tax_heads):
            ws.cell(row=6, column=2+i, value=val) # 3B
            ws.cell(row=6, column=6+i, value=val) # Ref
            ws.cell(row=6, column=10+i, value=val) # Diff
            
        # L0 headers (Row 5) - Repeat them to simulate what Pandas sees or what merged cells impl would look like
        # We must ensure L0 is "3B" for cols 2-5, "Reference" for 6-9, "Difference" for 10-13
        for c in range(2, 6): ws.cell(row=5, column=c, value="3B")
        for c in range(6, 10): ws.cell(row=5, column=c, value="Reference")
        for c in range(10, 14): ws.cell(row=5, column=c, value="Difference")
        
        # Data Row (Row 7)
        # 3B = 100, Ref = 120. Diff = 3B - Ref = 100 - 120 = -20
        # Wait, if Ref(120) > 3B(100), then Shortfall is 20.
        # Diff in Excel is usually 3B - Ref? Or Ref - 3B?
        # Parser logic: "Difference: Must have SHORT/DIFFERENCE".
        # Parser logic calculates: diff = 3B - 1. liability = max(1 - 3B, 0).
        
        data_3b = [gstr3b_val] * 4
        data_ref = [gstr1_val] * 4
        data_diff = [b - r for b, r in zip(data_3b, data_ref)]
        
        row_data = ["Apr"] + data_3b + data_ref + data_diff
        for c, val in enumerate(row_data, 1):
            ws.cell(row=7, column=c, value=val)
            
        wb.save(filename)
        
    def test_excel_success(self):
        """Scenario 1: Excel Success. GSTR-1 (120) > 3B (100). Liability = 20."""
        self.create_mock_excel(self.test_excel, gstr3b_val=100.0, gstr1_val=120.0)
        
        res = self.parser._parse_group_a_liability(
            self.test_excel, "Tax Liability", "Cat", "tmpl", []
        )
        
        self.assertIsNotNone(res)
        # 4 heads * 20 = 80 total shortfall
        self.assertEqual(res["total_shortfall"], 80.0)
        
        table = res["issue_table_data"]
        # Row indices: 0=3B, 1=1, 2=Diff, 3=Liab
        self.assertEqual(table["rows"][0]["col1"], 100) # 3B IGST
        self.assertEqual(table["rows"][1]["col1"], 120) # GSTR-1 IGST
        self.assertEqual(table["rows"][3]["col1"], 20)  # Liability IGST
        
    def test_excel_failure(self):
        """Scenario 2: Excel Malformed/Missing Sheet -> Returns None (Alert), No Fallback even if PDF provided."""
        # Create dummy excel without correct sheet
        wb = openpyxl.Workbook()
        wb.save(self.test_excel)
        
        # Mock PDFs existing
        with open(self.test_pdf_3b, 'w') as f: f.write("dummy")
        with open(self.test_pdf_1, 'w') as f: f.write("dummy")
        
        with patch('src.services.scrutiny_parser.parse_gstr3b_pdf_table_3_1_a') as m3b, \
             patch('src.services.scrutiny_parser.parse_gstr1_pdf_total_liability') as m1:
            # Should NOT be called
            res = self.parser._parse_group_a_liability(
                self.test_excel, "Tax Liability", "Cat", "tmpl", [],
                gstr3b_pdf_path=self.test_pdf_3b, gstr1_pdf_path=self.test_pdf_1
            )
            self.assertIsNone(res)
            m3b.assert_not_called()
            m1.assert_not_called()

    @patch('src.services.scrutiny_parser.parse_gstr3b_pdf_table_3_1_a')
    @patch('src.services.scrutiny_parser.parse_gstr1_pdf_total_liability')
    def test_pdf_priority_no_excel(self, m1, m3b):
        """Scenario 3: No Excel provided (file_path=None). Valid PDFs. Logic should parse PDFs."""
        # Mock return values
        m3b.return_value = {"igst": 100.0, "cgst": 100.0, "sgst": 100.0, "cess": 0.0}
        m1.return_value  = {"igst": 150.0, "cgst": 150.0, "sgst": 150.0, "cess": 0.0}
        
        # Mock file existence
        with open(self.test_pdf_3b, 'w') as f: f.write("dummy")
        with open(self.test_pdf_1, 'w') as f: f.write("dummy")
        
        res = self.parser._parse_group_a_liability(
            None, "Tax Liability", "Cat", "tmpl", [],
            gstr3b_pdf_path=self.test_pdf_3b, gstr1_pdf_path=self.test_pdf_1
        )
        
        self.assertIsNotNone(res)
        # Shortfall = 150 - 100 = 50 per head (3 heads) -> 150 total
        self.assertEqual(res["total_shortfall"], 150.0)
        table = res["issue_table_data"]
        self.assertEqual(table["rows"][1]["col1"], 150) # GSTR-1
        self.assertEqual(table["rows"][0]["col1"], 100) # 3B

    @patch('src.services.scrutiny_parser.parse_gstr3b_pdf_table_3_1_a')
    @patch('src.services.scrutiny_parser.parse_gstr1_pdf_total_liability')
    def test_mismatch_warning(self, m1, m3b):
        """Scenario 4: Excel + PDF Mismatch. Excel wins. Warning added."""
        self.create_mock_excel(self.test_excel, gstr3b_val=100.0, gstr1_val=100.0)
        
        # PDF says 200 (Mismatch)
        m3b.return_value = {"igst": 200.0, "cgst": 200.0, "sgst": 200.0, "cess": 0.0}
        m1.return_value  = {"igst": 100.0, "cgst": 100.0, "sgst": 100.0, "cess": 0.0}
        
        with open(self.test_pdf_3b, 'w') as f: f.write("dummy")
        with open(self.test_pdf_1, 'w') as f: f.write("dummy")
        
        res = self.parser._parse_group_a_liability(
            self.test_excel, "Tax Liability", "Cat", "tmpl", [],
            gstr3b_pdf_path=self.test_pdf_3b, gstr1_pdf_path=self.test_pdf_1
        )
        
        self.assertIsNotNone(res)
        # Excel data used -> 100 vs 100 -> 0 shortfall
        self.assertEqual(res["total_shortfall"], 0.0)
        
        # Check warnings
        self.assertIn("Warnings", res["status_msg"])
        self.assertTrue(any("GSTR-3B" in w and "200.0" in w for w in res["status_msg"].split("; ")))

if __name__ == '__main__':
    unittest.main()
