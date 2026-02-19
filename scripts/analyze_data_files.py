import fitz
import openpyxl
import os

GSTR3B_PDF = "GSTR3B_32AAMFM4610Q1Z0_032023.pdf"
GSTR2B_EXCEL = "062022_32AAMFM4610Q1Z0_GSTR2BQ_05012026.xlsx"

def inspect_pdf():
    print(f"\n--- Analyzing PDF: {GSTR3B_PDF} ---")
    if not os.path.exists(GSTR3B_PDF):
        print("PDF file not found.")
        return

    try:
        doc = fitz.open(GSTR3B_PDF)
        print(f"Pages: {len(doc)}")
        for i, page in enumerate(doc):
            text = page.get_text()
            print(f"Page {i+1} Extract (First 500 chars):\n{text[:500]}...")
            print("-" * 20)
    except Exception as e:
        print(f"Error reading PDF: {e}")

def inspect_excel():
    print(f"\n--- Analyzing Excel: {GSTR2B_EXCEL} ---")
    if not os.path.exists(GSTR2B_EXCEL):
        print("Excel file not found.")
        return

    try:
        wb = openpyxl.load_workbook(GSTR2B_EXCEL, read_only=True, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for sheet in wb.sheetnames:
             ws = wb[sheet]
             rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
             print(f"Sheet '{sheet}' First 5 rows:")
             for row in rows:
                 print(row)
             print("-" * 20)
    except Exception as e:
        print(f"Error reading Excel: {e}")

if __name__ == "__main__":
    inspect_pdf()
    inspect_excel()
