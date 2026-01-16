import pandas as pd
import openpyxl
import sys

# Force UTF-8 for output
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\manum\.gemini\antigravity\scratch\gst\032023_32AAMFM4610Q1Z0_GSTR2BQ_05012026.xlsx"

print(f"--- Scanning ALL rows for DEBIT NOTES in {file_path} ---")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    for sheet in ["B2B-CDNR", "B2B-CDNRA"]:
        if sheet in wb.sheetnames:
            print(f"\n--- Sheet: {sheet} ---")
            ws = wb[sheet]
            found_any = False
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                row_str = " ".join([str(x) for x in row if x]).lower()
                if "debit" in row_str:
                     print(f"Row {i+1}: Found DEBIT NOTE -> {row}")
                     found_any = True
            
            if not found_any:
                print("No Debit Notes found in this sheet.")

except Exception as e:
    print(f"Error: {e}")
