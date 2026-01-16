import pandas as pd
import openpyxl
import sys

# Force UTF-8 for output
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\manum\.gemini\antigravity\scratch\gst\032023_32AAMFM4610Q1Z0_GSTR2BQ_05012026.xlsx"

print(f"--- Inspecting Rows 28-35 in 'ITC Available' Sheet in {file_path} ---")

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if "ITC Available" in wb.sheetnames:
        ws = wb["ITC Available"]
        # 1-based index range inclusive
        for r_idx in range(28, 36):
            clean_row = []
            for cell in ws[r_idx]:
                 if cell.value is None:
                     clean_row.append("None")
                 else:
                     clean_row.append(str(cell.value).replace('\u20b9', 'Rs').strip())
            print(f"Row {r_idx}: {clean_row}")
    else:
        print("Sheet 'ITC Available' NOT FOUND.")

except Exception as e:
    print(f"Error: {e}")
