import openpyxl
import sys

EXCEL_FILE = "new sops.xlsx"

def analyze_excel_safe():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        print(f"Sheets found: {wb.sheetnames}")
        
        for sheet in wb.sheetnames:
            print(f"\n--- Analyzing Sheet: {sheet} ---")
            ws = wb[sheet]
            rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
            for row in rows:
                print(row)
            print("-" * 30)
            
    except Exception as e:
        print(f"Error reading Excel file: {e}")

if __name__ == "__main__":
    analyze_excel_safe()
