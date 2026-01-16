import pandas as pd
import openpyxl
import os
import sys

# Go up one level to project root to allow 'src' import
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from src.services.scrutiny_parser import ScrutinyParser

def create_mock_excel(filename):
    wb = openpyxl.Workbook()
    # Create "Tax Liability" sheet (case insensitive match in parser, excluding 'summary')
    ws = wb.active
    ws.title = "Tax Liability"
    
    # 1. Headers at B5, F5, J5 (Row 5, Cols 2, 6, 10)
    ws['B5'] = "Liability per 3B (Test)"
    ws['F5'] = "Liability per GSTR-1 (Test)"
    ws['J5'] = "Difference (Test)"
    
    # 2. DataFrame Headers at Row 4, 5 (Indices 4, 5 -> Excel Rows 5, 6)
    # Parser reads header=[4,5].
    # Let's populate some data starting Row 7 (Excel Row 8?).
    # Wait, Parser says: header=[4, 5].
    # Row 1-3 are ignored. Row 4 (Index 3) is Issue Name.
    ws['A4'] = "1. Test Issue Name"
    
    # Headers
    # L0: B4, C4... 
    # Let's mock the layout roughly so pandas can read it.
    # B4="3B", C4="3B"... F4="Ref"... J4="Diff"
    
    headers_l0 = ["Period", "3B", "3B", "3B", "3B", "Reference", "Reference", "Reference", "Reference", "Difference", "Difference", "Difference", "Difference"]
    headers_l1 = ["", "IGST", "CGST", "SGST", "CESS", "IGST", "CGST", "SGST", "CESS", "IGST", "CGST", "SGST", "CESS"]
    
    for c, val in enumerate(headers_l0, 1):
        ws.cell(row=5, column=c, value=val)
    for c, val in enumerate(headers_l1, 1):
        ws.cell(row=6, column=c, value=val)
        
    # Data Rows (Row 7+)
    data = [
        ["Apr", 100, 100, 100, 0, 120, 120, 120, 0, -20, -20, -20, 0], # Shortfall 20 each
        ["May", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    ]
    
    for r_idx, row_data in enumerate(data, 7):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
            
    wb.save(filename)
    wb.close()
    return filename

def test_extraction():
    parser = ScrutinyParser()
    filename = "test_tax_liability.xlsx"
    try:
        create_mock_excel(filename)
        print(f"Created {filename}")
        
        # Call the method
        res = parser._parse_group_a_liability(
            filename, 
            "Tax Liability", 
            "Test Category", 
            "test_template", 
            []
        )
        
        if not res:
            print("[FAIL] Parsing returned None")
            return
            
        print("[PASS] Parsing Successful")
        print(f"Total Shortfall: {res['total_shortfall']}")
        
        if "issue_table_data" in res:
            print("\n[PASS] Summary Table Found (issue_table_data):")
            st = res["issue_table_data"]
            print("Headers:", st["headers"])
            for row in st["rows"]:
                print(row)
                
            # validations
            rows = st["rows"]
            # Row 0 (3B) -> 100+0 = 100 per head
            # Row 1 (Ref) -> 120+0 = 120 per head
            # Row 2 (Diff) -> -20 per head? 
            # Wait, the parser logic accumulates totals.
            # Apr: 3B=100, Ref=120, Diff=-20
            # May: 0
            # Totals: 3B=100, Ref=120, Diff=-20
            
            # Row 0: Liability per 3B (Test), 100, 100, 100
            # Row 2: Difference (Test), -20, -20, -20
            
            assert rows[0]['col1'] == 100 # CGST
            assert rows[2]['col1'] == -20 # CGST Diff
            print("\n[PASS] Values Verified: Totals match inputs.")
            
            print("\n[PASS] Values Verified: Totals match inputs.")
            
        else:
            print("[FAIL] Summary Table (issue_table_data) Missing from Result")
            print(f"Keys found: {list(res.keys())}")
            
    except Exception as e:
        print(f"[ERROR] Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if os.path.exists(filename):
            os.remove(filename)

if __name__ == "__main__":
    test_extraction()
