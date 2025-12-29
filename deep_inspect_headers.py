import pandas as pd
import openpyxl
import os

target_file = "2022-23_32AABCL1984A1Z0_Tax liability and ITC comparison.xlsx"
sheet_name = "Tax liability"

if os.path.exists(target_file):
    print(f"--- Inspecting {target_file} [{sheet_name}] ---")
    try:
        # OpenPyXL Inspection
        wb = openpyxl.load_workbook(target_file, data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"B5: {ws['B5'].value}")
            print(f"F5: {ws['F5'].value}")
            print(f"J5: {ws['J5'].value}")
            
            print("-" * 20)
            for r in range(1, 10):
                row_vals = [c.value for c in ws[r]]
                print(f"Row {r}: {row_vals}")
        else:
            print(f"Sheet {sheet_name} not found in {wb.sheetnames}")
    except Exception as e:
        print(f"Error: {e}")
        
else:
    print("Target file not found")
