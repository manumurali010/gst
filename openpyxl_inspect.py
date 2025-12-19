import openpyxl

file_path = "2022-23_32AFWPD9794D1Z0_Tax liability and ITC comparison.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

sheets = ["ITC (Other than IMPG)", "ITC (IMPG)", "RCM_LIABILITY_ITC"]

for sheet_name in sheets:
    print(f"\n=== {sheet_name} ===")
    ws = wb[sheet_name]
    
    # Inspect Row 5 (1-based)
    print("Row 5 Cells:")
    for col in range(1, 15):
        val = ws.cell(row=5, column=col).value
        print(f"  Col {col} ({ws.cell(row=5, column=col).coordinate}): {val}")
        
    # Inspect Row 6 (1-based)
    print("Row 6 Cells:")
    for col in range(1, 15):
        val = ws.cell(row=6, column=col).value
        print(f"  Col {col} ({ws.cell(row=6, column=col).coordinate}): {val}")
        
    # Find Total Row
    total_row_idx = None
    for r in range(1, 50): # Check first 50 rows
        if str(ws.cell(row=r, column=1).value).strip().lower() == "total":
            total_row_idx = r
            break
            
    if total_row_idx:
        print(f"Total Row ({total_row_idx}):")
        for col in range(1, 20):
            val = ws.cell(row=total_row_idx, column=col).value
            print(f"  Col {col}: {val}")
    else:
        print("TOTAL row not found!")

wb.close()
