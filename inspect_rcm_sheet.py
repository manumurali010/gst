import pandas as pd
import openpyxl
import os

target_file = "2022-23_32AABCL1984A1Z0_Tax liability and ITC comparison.xlsx"
sheet_name = "RCM_LIABILITY_ITC" # As specified by user, verified in file listing previously

if os.path.exists(target_file):
    print(f"--- Inspecting {target_file} [{sheet_name}] ---")
    try:
        # OpenPyXL Inspection to find headers row
        wb = openpyxl.load_workbook(target_file, data_only=True)
        # Case insensitive sheet search
        real_sheet_name = next((s for s in wb.sheetnames if sheet_name.lower() in s.lower()), None)
        
        if real_sheet_name:
            ws = wb[real_sheet_name]
            print(f"Found Sheet: {real_sheet_name}")
            
            # Print first 10 rows to identify headers
            for r in range(1, 10):
                row_vals = [c.value for c in ws[r]]
                print(f"Row {r}: {row_vals}")
        else:
            print(f"Sheet {sheet_name} not found in {wb.sheetnames}")
            
    except Exception as e:
        print(f"Error: {e}")
else:
    print("Target file not found")
