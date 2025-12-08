import pandas as pd
import json
import re
from openpyxl import load_workbook

def parse_cell_reference(ref):
    """Convert Excel reference (e.g., 'C12') to (row_idx, col_idx)"""
    match = re.match(r"([A-Z]+)([0-9]+)", ref)
    if not match:
        return None
    col_str, row_str = match.groups()
    row = int(row_str) - 1 # 0-indexed
    
    # Convert column letters to index (A=0, B=1, ... AA=26)
    col = 0
    for char in col_str:
        col = col * 26 + (ord(char) - ord('A') + 1)
    col -= 1
    
    return row, col

def excel_formula_to_python(formula, cell_map):
    """
    Convert simple Excel formula to Python.
    formula: string starting with '='
    cell_map: dict mapping 'C12' -> 'var_r12_c3'
    """
    if not formula.startswith('='):
        return formula
        
    expr = formula[1:]
    
    # Replace cell references with variable names
    # Sort keys by length desc to avoid replacing 'AA1' with 'A1' sub-match
    sorted_refs = sorted(cell_map.keys(), key=len, reverse=True)
    
    for ref in sorted_refs:
        # Use regex to ensure we match full cell references (e.g. B12 not B1 in B12)
        # Pattern: lookbehind for non-word char, match ref, lookahead for non-word char
        pattern = r'(?<![A-Z])' + ref + r'(?![0-9])'
        expr = re.sub(pattern, f"v['{cell_map[ref]}']", expr)
        
    return expr

def import_issues(file_path, output_path):
    wb = load_workbook(file_path, data_only=False)
    
    issues = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sheet1': continue # Skip default/summary sheets if any
        
        ws = wb[sheet_name]
        
        # 1. Dynamic Metadata & Mapping Extraction
        # Scan Column A for keys
        metadata = {}
        mapping_refs = {}
        table_start_row = 12 # Default fallback
        
        # Scan first 20 rows for metadata and mapping
        for r in range(1, 25):
            key_cell = ws.cell(row=r, column=2) # Column B is Key
            val_cell = ws.cell(row=r, column=3) # Column C is Value
            
            key = str(key_cell.value).strip().upper() if key_cell.value else ""
            # print(f"Row {r}: Key='{key}', Val='{val_cell.value}'") # DEBUG
            
            # Metadata Keys
            if "ISSUE NAME" in key:
                metadata['issue_name'] = val_cell.value
            elif "BRIEF FACTS" in key:
                metadata['brief_facts'] = val_cell.value
            elif "SECTIONS VIOLATED" in key or "LEGAL PROVISIONS" in key:
                metadata['sections_violated'] = val_cell.value
            elif "CONCLUSION" in key:
                metadata['conclusion'] = val_cell.value
                
            # Mapping Keys
            elif "TAX LIABILITY CGST" in key:
                mapping_refs['tax_cgst'] = val_cell
            elif "TAX LIABILITY SGST" in key:
                mapping_refs['tax_sgst'] = val_cell
            elif "TAX LIABILITY IGST" in key:
                mapping_refs['tax_igst'] = val_cell
            elif "INTEREST" in key:
                mapping_refs['interest'] = val_cell
            elif "PENALTY" in key:
                mapping_refs['penalty'] = val_cell
                
            # Table Header Detection
            elif "TABLE HEADERS" in key:
                table_start_row = r

        issue_name = metadata.get('issue_name', sheet_name)
        brief_facts = metadata.get('brief_facts', '')
        sections_violated = metadata.get('sections_violated', '')
        conclusion = metadata.get('conclusion', '')

        # 3. Table Structure
        # ... (Existing table logic, but using found table_start_row)
        
        # ... (Cell mapping logic same as before)
        
        # We need to map every cell in the table to a variable name
        cell_map = {} # 'C15' -> 'cell_15_3'
        
        # Scan table rows
        current_row = table_start_row + 2 # Start of data (Skip Header and Subheader)
        max_row = ws.max_row
        
        grid_data = []
        
        for r in range(current_row, max_row + 1):
            row_data = []
            has_data = False
            for c in range(2, 10): # Scan columns 2 to 9 (Skip Col A)
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if val is not None:
                    has_data = True
                
                # Create variable name
                var_name = f"cell_{r}_{c}"
                cell_ref = cell.coordinate # e.g. 'B15'
                cell_map[cell_ref] = var_name
                
                cell_info = {
                    "ref": cell_ref,
                    "var": var_name,
                    "value": val,
                    "type": "input" # default
                }
                
                if isinstance(val, str) and val.startswith('='):
                    cell_info["type"] = "formula"
                    cell_info["formula"] = val
                elif val is None:
                    cell_info["type"] = "input" # Empty cells are inputs
                else:
                    cell_info["type"] = "static" # Label or fixed number
                    
                row_data.append(cell_info)
            
            if has_data:
                grid_data.append(row_data)
        
        # Now process formulas with the complete cell_map
        for row in grid_data:
            for cell in row:
                if cell["type"] == "formula":
                    cell["python_formula"] = excel_formula_to_python(cell["formula"], cell_map)
        
        # Resolve Mapping References using cell_map
        def resolve_mapping(cell):
            if not cell: return None
            if cell.value and str(cell.value).startswith('='):
                ref = str(cell.value)[1:] # Remove '='
                return cell_map.get(ref, None)
            return None

        issue = {
            "issue_id": sheet_name,
            "issue_name": issue_name,
            "templates": {
                "brief_facts": brief_facts,
                "grounds": "", # Not explicitly in Excel, maybe combine?
                "legal": sections_violated,
                "conclusion": conclusion
            },
            "grid_data": grid_data,
            "cell_map": cell_map,
            "tax_demand_mapping": {
                "tax_cgst": resolve_mapping(mapping_refs.get('tax_cgst')),
                "tax_sgst": resolve_mapping(mapping_refs.get('tax_sgst')),
                "tax_igst": resolve_mapping(mapping_refs.get('tax_igst')),
                "interest": resolve_mapping(mapping_refs.get('interest')),
                "penalty": resolve_mapping(mapping_refs.get('penalty'))
            }
        }
        
        issues.append(issue)
        
    # Save to JSON
    output = {
        "meta": {"version": "2.0", "source": "excel_import"},
        "issues": issues
    }
    
    with open(output_path, 'w') as f:
        json.dump(output, f, indent=4)
        
    print(f"Imported {len(issues)} issues to {output_path}")

if __name__ == "__main__":
    import_issues("data/ISSUE DB.xlsx", "data/issues_from_excel.json")
