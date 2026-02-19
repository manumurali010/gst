
import pandas as pd
import openpyxl
import re
import os
import logging
import sys

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class GSTR2BAnalyzer:
    def __init__(self, file_path):
        self.file_path = file_path
        self.use_light_parser = False
        self.wb = None
        
        if not os.path.exists(file_path):
            # This is critical, let it raise or handle gracefully?
            # Raising is fine if file missing.
            raise FileNotFoundError(f"GSTR-2B file not found: {file_path}")
            
        try:
            # Attempt standard load
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            # Fallback to Light Parser
            import logging
            logging.warning(f"GSTR2BAnalyzer: Failed to load with openpyxl ({e}). Switching to XLSXLight.")
            self.use_light_parser = True

    def validate_file(self, expected_gstin, expected_fy):
        """
        Validates GSTR-2B file structure and metadata (GSTIN, FY).
        Strict enforcement via regex on 'Read me' sheet.
        """
        # 1. Sheet Existence Check
        required_sheets = ["Read me", "ITC Available"]
        for sheet in required_sheets:
            if sheet not in self.wb.sheetnames:
                raise ValueError(f"Invalid GSTR-2B: Missing required sheet '{sheet}'")

        # 2. Metadata Validation (Full Text Scan)
        read_me_ws = self.wb["Read me"]
        read_me_text = ""
        for row in read_me_ws.iter_rows():
            for cell in row:
                if cell.value:
                    read_me_text += str(cell.value) + " "
        
        # Regex Patterns
        gstin_pattern = r'\d{2}[A-Z]{5}\d{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}'
        fy_pattern = r'\d{4}-\d{2}'

        gstin_match = re.search(gstin_pattern, read_me_text)
        fy_match = re.search(fy_pattern, read_me_text)

        if not gstin_match:
            raise ValueError("Validation Failed: Could not extract GSTIN from 'Read me'.")
        if not fy_match:
            raise ValueError("Validation Failed: Could not extract Financial Year from 'Read me'.")

        extracted_gstin = gstin_match.group(0).strip()
        extracted_fy = fy_match.group(0).strip()

        # Strict Matching
        if extracted_gstin != expected_gstin:
            raise ValueError(f"GSTIN Mismatch: Expected {expected_gstin}, Found {extracted_gstin}")
        
        # Normalize FY format if needed (e.g. 2022-2023 -> 2022-23)
        # But user rule says "Exact Match". Assuming inputs are normalized.
        if extracted_fy != expected_fy:
             raise ValueError(f"Financial Year Mismatch: Expected {expected_fy}, Found {extracted_fy}")
        
        logger.info(f"GSTR-2B Validation Successful: {extracted_gstin} | {extracted_fy}")
        return True

    def _extract_tax_heads_from_row(self, row_idx, sheet_df):
        """
        Extracts IGST, CGST, SGST, Cess from a specific row in 'ITC Available'.
        Strict Logic:
        - Skips first 3 cols.
        - Checks divisibility by 4.
        - Supports ONLY Monthly (1 block) or Quarterly (4 blocks).
        """
        # Read the row
        if row_idx >= len(sheet_df):
            return {'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'cess': 0.0}
            
        row_data = sheet_df.iloc[row_idx].tolist()
        
        # Skip metadata (Cols 0, 1, 2)
        # Assuming row_data includes all columns read by pandas
        # We need to filter for numeric columns. 
        # CAUTION: Pandas might read empty cells.
        
        # Strategy: Get strictly the potential tax columns starting from index 3
        potential_tax_cols = row_data[3:]
        
        # Filter to ensure we are dealing with the numeric structure
        # Implementation Detail: "Collect all numeric columns" 
        # But empty cells in Excel might be NaNs. We treat NaNs as 0.0 IF they are part of the structure.
        # But 'numeric_cols' implies checking the layout. 
        # In GSTR-2B, columns are fixed. 
        # A simpler robust way: Take ALL columns from index 3 onwards? 
        # But trailing empty columns might mess up the "% 4" check.
        # Let's clean trailing NaNs.
        
        cleaned_vals = []
        for x in potential_tax_cols:
             if pd.notna(x) and str(x).strip() != '':
                  try:
                       cleaned_vals.append(float(x))
                  except ValueError:
                       # Non-numeric in a tax field? Should probably not happen or be 0.
                       cleaned_vals.append(0.0) 
             else:
                  # Empty cell inside the block -> 0.0
                  # But trailing empty cells need to be careful.
                  # GSTR-2B usually has fixed width. 
                  # Let's capture strictly the columns that *should* be there. 
                  # But we don't know the exact width for Monthly/Quarterly dynamically unless we check.
                  pass
        
        # REVISION: To implement "count blocks" correctly, we should NOT just drop NaNs arbitrarily because 
        # a monthly file has 4 cols, a quarterly has 16.
        # If I drop a NaN from the middle, I shift the values.
        # So I must preserve structure.
        
        # Improved Strategy:
        # Determine the width of the row that contains data.
        # Count non-empty/numeric columns? No, that's risky. 
        # Look at the dataframe width?
        # Sheet scan usually reads 'N' columns.
        # Let's look at consecutive numeric-ish values from Col 3.
        
        vals = []
        for x in potential_tax_cols:
             val = 0.0
             try:
                  if pd.notna(x):
                       val = float(x)
             except:
                  val = 0.0
             vals.append(val)
        
        # Trim trailing zeros? No, Quarterly file might end with values.
        # But `vals` will contain ALL columns to the right. 
        # If there are extra empty columns, `len(vals)` is huge.
        # We need to find the 'Tax Block' boundary.
        
        # Explicit check: 
        # Monthly 2B typically has IGST, CGST, SGST, Cess (4 cols).
        # Quarterly has 4 blocks of 4 (16 cols).
        # If the sheet has extra columns (e.g. advisory), we need to exclude them.
        # The user rule: "Numeric tax columns after metadata".
        # This implies we count the columns that *are tax columns*.
        # Usually these are contiguous.
        
        # Let's count significant columns. 
        # Actually, standard 2B has advisory at the end? row 5 says "Advisory".
        # So we should take cols 3 to End-1? Or check headers?
        # CONSTRAINT: "DO NOT use header-based column resolution".
        
        # Heuristic-free approach:
        # GSTR-2B structure is fixed. 
        # Monthly: 4 columns.
        # Quarterly: 16 columns.
        # Check specific lengths.
        
        # Filter strictly: take the list of float values until we hit non-numeric or end?
        # Or just take len(vals).
        # If len(vals) is e.g. 5 (4 tax + 1 advisory text), we fail divisibility.
        # So we must strip non-numeric tail.
        
        # Strip tail
        while vals and (potential_tax_cols[len(vals)-1] is None or pd.isna(potential_tax_cols[len(vals)-1]) or isinstance(potential_tax_cols[len(vals)-1], str)):
             # Check if the value in potential_tax_cols correspondng to tail is string/empty
             # Wait, I converted to float in `vals`.
             # Let's check the source `potential_tax_cols`.
             src_val = potential_tax_cols[len(vals)-1]
             # If it looks like text (Advisory), drop it.
             # If it is numeric (0.0), keep it.
             is_text = False
             if isinstance(src_val, str):
                  try:
                       float(src_val)
                  except:
                       is_text = True
             
             if is_text or pd.isna(src_val) or src_val == '':
                  vals.pop()
             else:
                  break
        
        count = len(vals)
        
        # Strict Format Check
        if count == 0:
             # Empty row?
             return {'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'cess': 0.0}

        if count % 4 != 0:
             raise ValueError(f"GSTR-2B Format Error: Numeric column count {count} is not divisible by 4.")
        
        num_blocks = count // 4
        
        if num_blocks == 1:
             # Monthly
             block = vals[0:4]
        elif num_blocks == 4:
             # Quarterly (Jan, Feb, Mar, Total) -> Take Index 3 (Total)
             block = vals[12:16]
        else:
             raise ValueError(f"GSTR-2B Format Error: Unsupported block count {num_blocks}. Expected 1 (Monthly) or 4 (Quarterly).")
             
        return {'igst': block[0], 'cgst': block[1], 'sgst': block[2], 'cess': block[3]}


    def analyze_sop_3(self):
        """
        SOP-3: ISD Credit
        Logic: Net = Inward ISD - Credit Notes
        """
        df = self.wb["ITC Available"]
        # Convert to pandas for easier row search by index
        # To preserve strict index mapping, we load with header=None
        # But openpyxl is already loaded. Let's use internal parser or just iterate.
        # Using self.wb["ITC Available"].values to get list of lists
        
        rows = list(self.wb["ITC Available"].values)
        # Convert to dataframe for easier handling (optional, but good for slicing)
        import pandas as pd
        df = pd.DataFrame(rows)
        
        row_idx_inward = -1
        row_idx_credit = -1
        
        for idx, row in df.iterrows():
            col1_text = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
            
            # [SOP-3 FIX] Fuzzy Matching (Case-insensitive, whitespace tolerant)
            # Match "Inward" AND "ISD"
            if "inward" in col1_text and "isd" in col1_text:
                row_idx_inward = idx
                logger.info(f"[SOP-3] Found 'Inward Supplies from ISD' at Row {idx}")
                
            # Match "Credit Notes" AND "ISD" (excluding Amendment)
            if "isd" in col1_text and "credit notes" in col1_text and "amendment" not in col1_text:
                row_idx_credit = idx
                logger.info(f"[SOP-3] Found 'ISD Credit Notes' at Row {idx}")
                
        # Missing Row Behavior: Treat as zero, verify logic
        vals_inward = {'igst':0.0, 'cgst':0.0, 'sgst':0.0, 'cess':0.0}
        vals_credit = {'igst':0.0, 'cgst':0.0, 'sgst':0.0, 'cess':0.0}
        
        if row_idx_inward != -1:
             # Use strict block extraction (Ignores block count, looks for last block of 4)
             res = self._extract_tax_block_strict(df.iloc[row_idx_inward].tolist())
             if res: vals_inward = res
        else:
             logger.info("SOP-3 Info: 'Inward Supplies from ISD' row not found. Assuming 0.")
             
        if row_idx_credit != -1:
             res = self._extract_tax_block_strict(df.iloc[row_idx_credit].tolist())
             if res: vals_credit = res
        else:
             logger.info("SOP-3 Info: 'ISD - Credit notes' row not found. Assuming 0.")

        # Netting (Zero Floored)
        net_igst = max(0.0, vals_inward.get('igst', 0.0) - vals_credit.get('igst', 0.0))
        net_cgst = max(0.0, vals_inward.get('cgst', 0.0) - vals_credit.get('cgst', 0.0))
        net_sgst = max(0.0, vals_inward.get('sgst', 0.0) - vals_credit.get('sgst', 0.0))
        net_cess = max(0.0, vals_inward.get('cess', 0.0) - vals_credit.get('cess', 0.0))
        
        return {
            'status': 'pass',
            'igst': net_igst,
            'cgst': net_cgst,
            'sgst': net_sgst,
            'cess': net_cess
        }

    def _extract_tax_block_strict(self, row_values):
        """
        Canonical Algorithm for extracting tax blocks (IGST, CGST, SGST, Cess).
        
        Algorithm:
        1. Filter row for purely numeric values (float/int).
        2. Validate count % 4 == 0.
        3. If count == 0 -> None (Empty/Format Error).
        4. Select LAST block of 4 (Total).
        5. Return dict.
        """
        numerics = []
        for val in row_values:
            # Proper numeric check
            if val is not None and isinstance(val, (int, float)):
                 # Check for NaN if it's a float? 
                 # Standard float('nan') exists.
                 # If using openpyxl, it returns None for empty.
                 # If using pandas (legacy), it might return NaN, but we are removing pandas.
                 # If using XLSXLight, it returns "" (str).
                 # So check for int/float is safe for openpyxl values.
                 # But standard python float('nan') is instance of float.
                 import math
                 if isinstance(val, float) and math.isnan(val): continue
                 numerics.append(float(val))
            elif isinstance(val, str):
                 # Try convert strict
                 try:
                     # Remove commas if any (though usually openpyxl gives float)
                     v_clean = val.replace(',', '').strip()
                     if v_clean:
                        numerics.append(float(v_clean))
                 except: pass
        
        count = len(numerics)
        
        if count == 0:
            return None # No data found
            
        if count % 4 != 0:
            logger.warning(f"GSTR-2B Strict Parse: Numeric count {count} not divisible by 4. Values: {numerics}")
            return None # Format Error
            
        # Select Last Block
        block = numerics[-4:]
        
        return {
            'igst': block[0],
            'cgst': block[1],
            'sgst': block[2],
            'cess': block[3]
        }

    def get_isd_raw_data(self):
        """
        Extracts raw tax heads for 'Inward supplies from ISD' row.
        """
        rows = list(self.wb["ITC Available"].values)
        import pandas as pd
        df = pd.DataFrame(rows)
        
        row_idx_inward = -1
        for idx, row in df.iterrows():
            col1_text = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
            if "inward supplies from isd" in col1_text:
                row_idx_inward = idx
                break
        
        if row_idx_inward != -1:
             # UPGRADE: Use strict extractor (Block Count Safe)
             row_vals = df.iloc[row_idx_inward].tolist()
             # Skip potential metadata calls if necessary, but strictly finding last block
             # _extract_tax_block_strict filters numerics and takes last 4.
             return self._extract_tax_block_strict(row_vals)
        else:
             return None

    def get_all_other_itc_raw_data(self):
        """
        Extracts raw tax values for 'All other ITC' from 'ITC Available' sheet.
        STRICT HARDENED LOGIC:
        - Uses Summary Row ONLY.
        - No Detail Sheet Iteration (Double Counting/Bug Risk).
        - Robust Row Matching.
        - Canonical Column Extraction.
        """
        try:
            if "ITC Available" not in self.wb.sheetnames:
                return None
            
            rows = list(self.wb["ITC Available"].values)
            import pandas as pd
            import re
            df = pd.DataFrame(rows)
            
            # Find the specific rows
            target_row_vals = None
            credit_note_row_vals = None
            credit_note_amend_row_vals = None
            
            for idx, row in df.iterrows():
                # Normalize row string for search
                # Join only string/text columns to identify row label
                # Index 1 usually contains Description
                
                # Robust Pattern Matching
                # We prioritize checking specific columns if we know them, but fallback to row-scan
                # usually Col 1 (B) has description.
                
                row_text_cells = [str(x).lower().strip() for x in row.values if isinstance(x, str)]
                row_str = " ".join(row_text_cells)
                
                # Markers: "all other itc" AND "registered persons"
                # Remove punctuation for matching
                row_clean = re.sub(r'[^\w\s]', '', row_str) 
                
                
                if "all other itc" in row_clean and "registered persons" in row_clean:
                    target_row_vals = row.tolist()
                    logger.info(f"SOP-4: Found Summary Row (Gross) at Index {idx}")
                    # Keep searching? No, we need other rows, so continue. But we found primary.
                    # Actually, we need to find Credit note rows too.
                
                # Check for Credit Notes Summary Rows
                # Row 30: "B2B - Credit notes"
                # Row 31: "B2B - Credit notes (Amendment)"
                
                # Exclude "Reverse charge" because Row 32 "B2B - Credit notes (Reverse charge)" also matches "B2B"+"Credit Notes"
                if "b2b" in row_clean and "credit notes" in row_clean and "amendment" not in row_clean and "reverse charge" not in row_clean:
                    # Original Credit Notes
                    credit_note_row_vals = row.tolist()
                    logger.info(f"SOP-4: Found Credit Note Row at Index {idx}")
                    
                if "b2b" in row_clean and "credit notes" in row_clean and "amendment" in row_clean and "reverse charge" not in row_clean:
                    # Amended Credit Notes
                    credit_note_amend_row_vals = row.tolist()
                    logger.info(f"SOP-4: Found Credit Note Amendment Row at Index {idx}")
            
            if target_row_vals is None: 
                logger.warning("SOP-4: 'All Other ITC' Summary Row NOT found.")
                return None
                
            # Extract Numeric Blocks
            # 1. Gross ITC (All Other)
            res = self._extract_tax_block_strict(target_row_vals)
            if res is None: return None
            
            # 2. Credit Notes (Original)
            cn_res = None
            if credit_note_row_vals:
                 cn_res = self._extract_tax_block_strict(credit_note_row_vals)
            
            # 3. Credit Notes (Amendment)
            cna_res = None
            if credit_note_amend_row_vals:
                 cna_res = self._extract_tax_block_strict(credit_note_amend_row_vals)
            
            # Netting: Gross - (CN + CN_Amend)
            # Note: 2B Summary sheet usually has positive values for Credit Notes row.
            # We subtract them from ITC available.
            
            if cn_res:
                logger.info(f"SOP-4: Subtracting Credit Notes: {cn_res}")
                for k in res: res[k] -= cn_res.get(k, 0.0)
                
            if cna_res:
                logger.info(f"SOP-4: Subtracting Credit Note Amendments: {cna_res}")
                for k in res: res[k] -= cna_res.get(k, 0.0)
            
            logger.info(f"SOP-4: Final Net ITC for file: {res}")
            
            # Zero Floor (Safety)
            for k in res:
                if res[k] < 0: res[k] = 0.0
                
            return res

        except Exception as e:
            logger.error(f"Error extracting All Other ITC raw data: {e}")
            return None

    def analyze_sop_10(self):
        """
        SOP-10: Import of Goods (IMPG + IMPGSEZ)
        Returns aggregated IGST.
        """
        try:
            if "ITC Available" not in self.wb.sheetnames:
                return {'status': 'info', 'reason': 'ITC Available sheet missing', 'igst': 0.0}

            rows = list(self.wb["ITC Available"].values)
            import pandas as pd
            df = pd.DataFrame(rows)
            
            row_idx_impg = -1
            row_idx_sez = -1
            row_idx_header = -1
            
            # [SOP-10 DIAG] Deep Inspection of Import Rows
            for idx, row in df.iterrows():
                row_text_parts = [str(x).lower().strip() for x in row.values if isinstance(x, str)]
                row_text = " ".join(row_text_parts)
                
                # Check for ANY import related keywords for diagnostics
                if "import" in row_text or "impg" in row_text:
                     vals = self._extract_tax_block_strict(row.tolist())
                     clean_text = (row_text[:50] + '...') if len(row_text) > 50 else row_text
                     logger.warning(f"[SOP-10 DIAG] Row {idx} VAL_CHECK: {bool(vals)} | '{clean_text}'")
                     if vals:
                         logger.warning(f"[SOP-10 DIAG] Row {idx} VALUES: {vals}")

                # 1. Detect Consolidated Header (IV. Import of goods...)
                # Strict check: "iv" + "import of goods"
                if "iv" in row_text and "import of goods" in row_text:
                    row_idx_header = idx

                if "import of goods" in row_text and "overseas" in row_text:
                    row_idx_impg = idx
                if "import of goods" in row_text and "sez" in row_text:
                    row_idx_sez = idx
            
            vals_impg = {'igst':0.0}
            vals_sez = {'igst':0.0}
            vals_header = None
            
            found_header = False
            found_components = False
            
            # Extract Header
            if row_idx_header != -1:
                res = self._extract_tax_block_strict(df.iloc[row_idx_header].tolist())
                if res and res.get('igst', 0) >= 0: # Accept 0, but must be valid struct
                    vals_header = res
                    found_header = True
            
            # Extract Components (Fallback/Check)
            if row_idx_impg != -1:
                 res = self._extract_tax_block_strict(df.iloc[row_idx_impg].tolist())
                 if res: 
                    vals_impg = res
                    found_components = True
            
            if row_idx_sez != -1:
                 res = self._extract_tax_block_strict(df.iloc[row_idx_sez].tolist())
                 if res:
                    vals_sez = res
                    found_components = True
            
            # Priority Logic
            if found_header:
                logger.info(f"[SOP-10 FIX] Using Consolidated Header Row {row_idx_header}: {vals_header}")
                if found_components:
                    logger.info("[SOP-10 FIX] Consolidated Header preferred over components.")
                
                return {
                    'status': 'pass',
                    'igst': float(vals_header.get('igst', 0.0))
                }
            
            # Fallback Logic (Component Sum)
            if not found_components:
                 return {'status': 'info', 'reason': 'Import data not found', 'igst': 0.0}
                 
            total_igst = vals_impg.get('igst', 0) + vals_sez.get('igst', 0)
            logger.warning(f"[SOP-10 INFO] Fallback to component sum (Missing Amendments risk): {total_igst}")
            
            return {
                'status': 'pass',
                'igst': float(total_igst)
            }
        except Exception as e:
            logger.error(f"Error in SOP-10 Analysis: {e}")
            return {'status': 'info', 'reason': str(e), 'igst': 0.0}

    # ==========================================
    # New Parsers for SOP 13-16 (RCM/Cash/Interest)
    # ==========================================

    def get_rcm_inward_supplies(self):
        """
        Extracts 'Inward Supplies Liable for Reverse Charge' from 'ITC Available'.
        Returns dict with int values {igst, cgst, sgst, cess}.
        Safeguards:
        - Matches "reverse" AND "inward" (case-insensitive)
        - Excludes "credit", "amendment", "details"
        - Prefers consolidated summary row
        - Logs warning if multiple candidate rows found
        """
        try:
            rows = []
            if self.use_light_parser or self.wb is None:
                 from src.utils.xlsx_light import XLSXLight
                 rows = XLSXLight.read_sheet(self.file_path, "ITC Available")
                 if not rows: return None
            else:
                 if "ITC Available" not in self.wb.sheetnames: return None
                 rows = list(self.wb["ITC Available"].values)
            
            candidate_rows = []
            
            for idx, row in enumerate(rows):
                if not row: continue
                row_list = list(row)
                
                # Build text for matching
                row_text_parts = [str(x).lower().strip() for x in row_list if x is not None and str(x).strip()]
                row_text = " ".join(row_text_parts).replace(",", "")
                
                # Flexible Matching Logic
                if "reverse" in row_text and "inward" in row_text:
                    # Exclusion Criteria
                    # "credit" matches "Input Tax Credit" in advisory text, so use "credit note"
                    if any(x in row_text for x in ["credit note", "amendment", "details", "invoice"]):
                        continue
                        
                    vals = self._extract_tax_block_strict(row_list)
                    if vals:
                        candidate_rows.append((idx, vals, row_text))
                        logger.info(f"[RCM DETECT] Candidate Row {idx}: {vals} | Text: {row_text[:50]}...")

            if not candidate_rows:
                logger.warning(f"[RCM DETECT] No suitable RCM Inward Supply rows found in {os.path.basename(self.file_path)}")
                return None
                
            # Selection Logic - Reverted to 'Pick Last' to avoid double counting Totals
            # (User confirmed the last block is usually the total)
            if len(candidate_rows) > 1:
                logger.warning(f"[RCM DETECT] Multiple RCM rows found ({len(candidate_rows)}). Selecting the LAST one as per standard format.")
                for i_r, v_r, t_r in candidate_rows:
                     logger.debug(f"  Candidate {i_r}: {v_r}")
            
            # Default: Pick the last one (often Total in quarterly)
            selected_idx, selected_vals, selected_text = candidate_rows[-1]
            logger.info(f"[RCM DETECT] Selected Row {selected_idx} (Final): {selected_vals}")
            
            return {k: int(round(v)) for k, v in selected_vals.items()}

        except Exception as e:
            logger.error(f"Error extracting RCM Inward Supplies: {e}")
            return None

    def get_rcm_credit_notes(self):
        """
        Extracts 'Credit Notes' (Reverse Charge) from 'ITC Available'.
        Target rows: 
          - "B2B - Credit Notes (Reverse Charge)"
          - "B2B - Credit Notes (Reverse Charge) (Amendment)"
        Returns dict with int values.
        """
        try:
            rows = []
            if self.use_light_parser or self.wb is None:
                 from src.utils.xlsx_light import XLSXLight
                 rows = XLSXLight.read_sheet(self.file_path, "ITC Available")
                 if not rows: return None
            else:
                 if "ITC Available" not in self.wb.sheetnames: return None
                 rows = list(self.wb["ITC Available"].values)
            
            total_cn = {'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'cess': 0.0}
            found_any_row = False
            
            for idx, row in enumerate(rows):
                if not row: continue
                row_list = list(row)
                
                # Normalize text
                row_text_parts = [str(x).lower().strip() for x in row_list if x is not None and str(x).strip()]
                row_text = " ".join(row_text_parts).replace(",", "")
                
                # Strict Match Logic for RCM Credit Notes
                # Must contain: "credit", "reverse", "b2b"
                # Must NOT contain: "others", "net-off" (advisory text check)
                
                if "reverse" in row_text and "credit" in row_text:
                     # Check for specific B2B context to avoid "Others" row with advisory text
                     if "b2b" in row_text:
                         vals = self._extract_tax_block_strict(row_list)
                         if vals:
                             found_any_row = True
                             logger.info(f"[RCM CN MATCH] Row {idx}: {vals} | Text: {row_text[:50]}...")
                             for k, v in vals.items():
                                 # Sum values exactly (preserve sign)
                                 total_cn[k] += v
                     else:
                         # Log ignored rows helpful for debugging
                         if "b2b" not in row_text and "note" in row_text:
                             logger.debug(f"[RCM CN SKIP] Row {idx} ignored (missing 'b2b'): {row_text[:50]}...")

            if not found_any_row:
                logger.warning(f"[RCM DETECT] No 'B2B - Credit Notes (Reverse Charge)' rows found in {os.path.basename(self.file_path)}")
                return None
            
            logger.info(f"[RCM CN TOTAL] Final Sum: {total_cn}")
            return {k: int(round(v)) for k, v in total_cn.items()}

        except Exception as e:
            logger.error(f"Error extracting RCM Credit Notes: {e}")
            return None
