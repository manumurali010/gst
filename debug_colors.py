import openpyxl

file_path = "c:\\Users\\manum\\.gemini\\antigravity\\scratch\\GST_Adjudication_System\\2022-23_32AFWPD9794D1Z0_Tax liability and ITC comparison.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if "ITC (Other" in s), None)
    ws = wb[sheet_name]
    
    print("Scanning for Colored Cells...")
    found = False
    for row in ws.iter_rows(min_row=8, max_row=100, max_col=20):
        for cell in row:
            if cell.value is not None:
                c = cell.font.color
                # Check if it has a color object
                if c:
                    # Filter out standard Black (Theme 1)
                    if c.type == 'theme' and c.theme != 1:
                        print(f"Row {cell.row} Col {cell.column}: Val={cell.value} | Theme={c.theme} | Tint={c.tint}")
                        found = True
                    # Check RGB if available
                    elif c.type == 'rgb' and str(c.rgb).upper() not in ['FF000000', '000000']:
                        print(f"Row {cell.row} Col {cell.column}: Val={cell.value} | RGB={c.rgb}")
                        found = True
                        
    if not found:
        print("No non-standard colored cells found in first 100 rows.")

except Exception as e:
    print(f"Error: {e}")
