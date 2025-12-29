import pandas as pd
import openpyxl
import os

file_path = "2022-23_32AABCL1984A1Z0_Tax liability and ITC comparison.xlsx"
sheet_name = "RCM_LIABILITY_ITC"

def debug_columns():
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    print(f"Analyzing {file_path} for '{sheet_name}'...")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    real_sheet_name = next((s for s in wb.sheetnames if sheet_name.lower() in s.lower()), None)
    
    if not real_sheet_name:
        print("Sheet not found.")
        return
        
    ws = wb[real_sheet_name]
    
    # Read Headers
    df = pd.read_excel(file_path, sheet_name=real_sheet_name, header=[4, 5])

    last_valid_l0 = ""

    with open("rcm_cols.txt", "w") as f:
        f.write(f"Analyzing {file_path} for '{sheet_name}'...\n")
        f.write(f"B5: {ws['B5'].value}\n")
        f.write(f"F5: {ws['F5'].value}\n")
        f.write(f"J5: {ws['J5'].value}\n")

        
        f.write("\n--- RCM Column Mapping ---\n")
        for i, col in enumerate(df.columns):
            l0 = str(col[0]).strip()
            if "Unnamed" in l0 or l0 == "nan":
                l0 = last_valid_l0
            else:
                last_valid_l0 = l0
                
            full = f"{l0} {col[1]}".upper()
            f.write(f"{i:<3} | {full}\n")
            
    print("Done writing to rcm_cols.txt")

if __name__ == "__main__":
    debug_columns()
